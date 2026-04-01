"""
run_clean.py —— 数据清洗：Excel → 清洗后 Excel

清洗步骤（与 run_pipeline 的校验逻辑一致）：
  1. 移除 input / output 列中的空值行
  2. 移除 output 列中不在 valid_labels 列表内的非法标签行
  3. 基于 input 列去重（保留第一条）

输出：
  data/processed/cleaned_<timestamp>.xlsx
  列宽可通过 config.yaml 或命令行参数控制

典型用法：
  # 使用 config.yaml 默认配置
  uv run python scripts/run_clean.py --input data/raw/sample.xlsx

  # 自定义 input 列宽（字符数）
  uv run python scripts/run_clean.py --input data/raw/sample.xlsx --input-col-width 80

  # 指定输出路径
  uv run python scripts/run_clean.py --input data/raw/sample.xlsx --output data/processed/my_clean.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import get_settings
from src.loader import ExcelLoader
from src.validator import DataValidator

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("run_clean")

# ── 默认列宽（字符数）──
_DEFAULT_INPUT_COL_WIDTH  = 60
_DEFAULT_OUTPUT_COL_WIDTH = 16


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="data-master：Excel 数据清洗（输出清洗后的 Excel）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="输入 Excel 文件路径",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 Excel 路径（默认：data/processed/cleaned_<时间戳>.xlsx）",
    )
    parser.add_argument(
        "--input-col-width",
        type=int,
        default=None,
        metavar="INT",
        help=f"input 列宽（字符数，默认 {_DEFAULT_INPUT_COL_WIDTH}）",
    )
    parser.add_argument(
        "--output-col-width",
        type=int,
        default=None,
        metavar="INT",
        help=f"output 列宽（字符数，默认 {_DEFAULT_OUTPUT_COL_WIDTH}）",
    )
    return parser.parse_args()


def _get_col_widths(cfg, input_override: int | None, output_override: int | None) -> dict[str, int]:
    """从 config.yaml clean 节读取列宽，命令行参数优先级更高。"""
    import yaml  # noqa: PLC0415
    raw: dict = {}
    config_path = _PROJECT_ROOT / "config.yaml"
    if config_path.exists():
        with config_path.open("r", encoding="utf-8") as f:
            raw = yaml.safe_load(f) or {}

    clean_cfg = raw.get("clean", {}).get("col_widths", {})
    input_w  = input_override  or clean_cfg.get("input",  _DEFAULT_INPUT_COL_WIDTH)
    output_w = output_override or clean_cfg.get("output", _DEFAULT_OUTPUT_COL_WIDTH)
    return {cfg.input_col: int(input_w), cfg.output_col: int(output_w)}


def save_to_excel(
    df,
    output_path: Path,
    col_widths: dict[str, int],
) -> None:
    """将 DataFrame 写入 Excel，并应用列宽、标题样式。

    Args:
        df:          待写入的 DataFrame。
        output_path: 输出文件路径（.xlsx）。
        col_widths:  {列名: 列宽（字符数）} 映射。
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cleaned"

    # ── 标题行样式 ──
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── 写标题行 ──
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    # ── 数据行样式 ──
    data_align = Alignment(vertical="top", wrap_text=True)

    # ── 写数据行 ──
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align

    # ── 设置列宽 ──
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        width  = col_widths.get(col_name, 20)
        ws.column_dimensions[letter].width = width

    # ── 冻结首行（方便滚动查看）──
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Excel 已保存：{output_path}  ({len(df)} 行)")


def main() -> None:
    args = parse_args()
    cfg  = get_settings()

    # ── 确定输出路径 ──
    if args.output:
        output_path = args.output
    else:
        output_path = (
            cfg.data_processed_dir
            / f"cleaned_{cfg.run_timestamp}.xlsx"
        )

    # ── 读取列宽配置 ──
    col_widths = _get_col_widths(cfg, args.input_col_width, args.output_col_width)

    print("\n🧹 run_clean：数据清洗")
    print(f"   输入：{args.input.resolve()}")
    print(f"   输出：{output_path.resolve()}")
    print(f"   列宽：input={col_widths[cfg.input_col]}  output={col_widths[cfg.output_col]}\n")

    # ── Step 1：加载（支持多 sheet，自动忽略多余列）──
    loader = ExcelLoader(cfg)
    df_raw = loader.load(args.input)
    print(f"📂 加载完成：{len(df_raw)} 条\n")

    # ── Step 2：清洗 ──
    validator = DataValidator(cfg)
    result = validator.validate(df_raw)
    print(result.summary())

    if len(result.cleaned_df) == 0:
        print("\n❌ 清洗后数据为空，请检查原始数据。")
        sys.exit(1)

    # ── Step 3：写出 Excel ──
    print(f"\n💾 写出清洗后数据...")
    save_to_excel(result.cleaned_df, output_path, col_widths)

    removed = len(df_raw) - len(result.cleaned_df)
    print(
        f"\n✅ 清洗完成：{len(df_raw)} 条 → {len(result.cleaned_df)} 条"
        f"（移除 {removed} 条无效/重复数据）"
    )
    print(f"   输出文件：{output_path.resolve()}")


if __name__ == "__main__":
    main()
